from openpyxl import load_workbook

def translate(target_value):
    # 엑셀 파일 불러오기
    sheet_name = 'GRI Index 번역'
    file_path = 'TM_noRAG/GRI_index_translate.xlsx'
    workbook = load_workbook(file_path)
    # 시트 선택
    sheet = workbook[sheet_name]

    # 각 셀을 순회하면서 값을 찾음
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == target_value:
                # 찾은 셀의 오른쪽 셀 값 반환
                right_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                return right_cell.value  # 오른쪽 셀의 값 반환

    # 찾지 못한 경우 None 반환
    return None

# 사용 예시
# target_value = 'GRI 201-1'
# right_cell_value = translate(target_value)
# print(right_cell_value)
